from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, List
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from api.core.database import get_db

router = APIRouter()

def _fix_id(doc: dict) -> dict:
    if doc and "_id" in doc:
        doc["_id"] = str(doc["_id"])
    return doc

async def _get_filtered_publications(filters: dict) -> list:
    """Fetch all filtered publications (no pagination)"""
    db = get_db()
    query = {}

    if filters.get("year"):
        query["year"] = filters["year"]
    elif filters.get("year_from") or filters.get("year_to"):
        year_filter = {}
        if filters.get("year_from"):
            year_filter["$gte"] = filters["year_from"]
        if filters.get("year_to"):
            year_filter["$lte"] = filters["year_to"]
        query["year"] = year_filter

    if filters.get("pub_type"):
        query["pub_type"] = filters["pub_type"]
    else:
        # Exclude 'unknown' pub_type to match frontend behavior
        query["pub_type"] = {"$in": ["journal", "conference", "book"]}

    if filters.get("author_id"):
        query["author_id"] = filters["author_id"]
    elif filters.get("author_name"):
        query["author_name"] = {"$regex": filters["author_name"], "$options": "i"}

    if filters.get("min_citations") is not None or filters.get("max_citations") is not None:
        cit_filter = {}
        if filters.get("min_citations") is not None:
            cit_filter["$gte"] = filters["min_citations"]
        if filters.get("max_citations") is not None:
            cit_filter["$lte"] = filters["max_citations"]
        query["cited_by"] = cit_filter

    allowed_sorts = {"cited_by", "year", "title", "author_name"}
    sort_by = filters.get("sort_by", "cited_by")
    if sort_by not in allowed_sorts:
        sort_by = "cited_by"
    sort_dir = -1 if filters.get("order", "desc") == "desc" else 1

    pubs = (
        await db.publications.find(query)
        .sort([("author_name", 1), (sort_by, sort_dir)])
        .to_list(None)  # Fetch all results
    )

    return [_fix_id(p) for p in pubs]

FIELD_MAP = {
    "all_authors": {"label": "Authors", "width": 35, "pdf_width": 1.5 * inch},
    "title": {"label": "Title", "width": 40, "pdf_width": 3.0 * inch},
    "venue": {"label": "Journal/Conference", "width": 25, "pdf_width": 1.2 * inch},
    "year": {"label": "Year", "width": 10, "pdf_width": 0.6 * inch},
    "pub_type": {"label": "Type", "width": 15, "pdf_width": 0.8 * inch},
    "link": {"label": "Link", "width": 30, "pdf_width": 1.2 * inch},
}

@router.get("/excel")
async def export_to_excel(
    author_id: Optional[str] = Query(None),
    author_name: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    pub_type: Optional[str] = Query(None),
    min_citations: Optional[int] = Query(None),
    max_citations: Optional[int] = Query(None),
    year_from: Optional[int] = Query(None),
    year_to: Optional[int] = Query(None),
    sort_by: str = Query("cited_by"),
    order: str = Query("desc"),
    fields: List[str] = Query(None)
):
    """
    Export filtered publications to Excel
    """
    try:
        filters = {
            "author_id": author_id,
            "author_name": author_name,
            "year": year,
            "pub_type": pub_type,
            "min_citations": min_citations,
            "max_citations": max_citations,
            "year_from": year_from,
            "year_to": year_to,
            "sort_by": sort_by,
            "order": order,
        }
        
        publications = await _get_filtered_publications(filters)
        
        if not publications:
            raise HTTPException(status_code=404, detail="No publications found matching the filters")
        
        if not fields:
            fields = ["all_authors", "title", "venue", "year", "pub_type", "link"]
            
        selected_fields = [f for f in fields if f in FIELD_MAP]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Publications"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num, field_id in enumerate(selected_fields, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = FIELD_MAP[field_id]["label"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = FIELD_MAP[field_id]["width"]
            
        for row_num, pub in enumerate(publications, 2):
            for col_num, field_id in enumerate(selected_fields, 1):
                if field_id == "all_authors":
                    val = pub.get("all_authors") or pub.get("author_name", "")
                else:
                    val = pub.get(field_id, "")
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(val) if val is not None else ""
                if field_id == "title":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"publications_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            iter([excel_buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")


@router.get("/pdf")
async def export_to_pdf(
    author_id: Optional[str] = Query(None),
    author_name: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    pub_type: Optional[str] = Query(None),
    min_citations: Optional[int] = Query(None),
    max_citations: Optional[int] = Query(None),
    year_from: Optional[int] = Query(None),
    year_to: Optional[int] = Query(None),
    sort_by: str = Query("cited_by"),
    order: str = Query("desc"),
    fields: List[str] = Query(None)
):
    """
    Export filtered publications to PDF
    """
    try:
        filters = {
            "author_id": author_id,
            "author_name": author_name,
            "year": year,
            "pub_type": pub_type,
            "min_citations": min_citations,
            "max_citations": max_citations,
            "year_from": year_from,
            "year_to": year_to,
            "sort_by": sort_by,
            "order": order,
        }
        
        publications = await _get_filtered_publications(filters)
        
        if not publications:
            raise HTTPException(status_code=404, detail="No publications found matching the filters")
            
        if not fields:
            fields = ["all_authors", "title", "venue", "year", "pub_type"]
            
        selected_fields = [f for f in fields if f in FIELD_MAP]
        
        pdf_buffer = io.BytesIO()
        # Use landscape for more columns if needed, but letter is fine for now
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=1
        )
        
        story.append(Paragraph("DCSE Faculty Publications Export", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=0
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.whitesmoke,
            fontName='Helvetica-Bold',
            alignment=0
        )
        
        # Headers
        header_row = [Paragraph(FIELD_MAP[f]["label"], header_style) for f in selected_fields]
        table_data = [header_row]
        
        # Data
        for pub in publications:
            row = []
            for f in selected_fields:
                if f == "all_authors":
                    raw_val = pub.get("all_authors") or pub.get("author_name", "")
                else:
                    raw_val = pub.get(f, "")
                val = str(raw_val) if raw_val is not None else ""
                row.append(Paragraph(val, cell_style))
            table_data.append(row)
            
        col_widths = [FIELD_MAP[f]["pdf_width"] for f in selected_fields]
        
        # If total width exceeds page, scale down proportionally
        total_width = sum(col_widths)
        available_width = letter[0] - doc.leftMargin - doc.rightMargin
        if total_width > available_width:
            scale_factor = available_width / total_width
            col_widths = [w * scale_factor for w in col_widths]
        
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>Total Publications: {len(publications)}</b>", styles['Normal']))
        
        doc.build(story)
        pdf_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"publications_export_{timestamp}.pdf"
        
        return StreamingResponse(
            iter([pdf_buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF file: {str(e)}")
